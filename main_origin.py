import openai
import os
import json
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import docx
import docx2txt
from PyPDF2 import PdfReader
import textract

# 固定 OpenAI API 密鑰
OPENAI_API_KEY = 'sk-proj-AtL42GrFHO5LI48oQNL8T3BlbkFJEDyALsIfyjCWJN72OQkm'

def analyze_file(file_path):
    file_extension = os.path.splitext(file_path)[1].lower()
    try:
        if file_extension in ['.pdf', '.doc', '.docx']:
            text = textract.process(file_path).decode('utf-8')
            return text
        else:
            print(f"不支持的文件類型：{file_extension}")
            return None
    except Exception as e:
        print(f"處理文件 {file_path} 時出錯: {str(e)}")
        return None

def calculate_credits(duration_minutes, credit_type):
    if credit_type == "甲類":
        return math.ceil(duration_minutes / 50)
    else:  # 乙類
        return round(duration_minutes / 50 * 0.5, 1)

def get_gpt4_json_response(client, prompt):
    response = client.chat.completions.create(
        model="gpt-4o",
        messages=[
            {"role": "system", "content": "你是一個專業的文件分析助手。請以JSON格式返回分析結果。"},
            {"role": "user", "content": prompt}
        ],
        response_format={"type": "json_object"}
    )
    return json.loads(response.choices[0].message.content)

def calculate_duration(start_time, end_time):
    start = datetime.strptime(start_time, "%H:%M")
    end = datetime.strptime(end_time, "%H:%M")
    if end < start:
        end += timedelta(days=1)  # 處理跨午夜的情況
    duration = end - start
    return int(duration.total_seconds() / 60)  # 返回分鐘數

def process_topics(topics):
    processed_topics = []
    for topic in topics:
        if '（包含 QA）' in topic['topic']:
            start_time, end_time = topic['time'].split('-')
            duration = calculate_duration(start_time.strip(), end_time.strip())
            topic['duration'] = duration
        processed_topics.append(topic)
    return processed_topics

def write_to_excel(all_results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "課程分析結果"

    # 設置標題樣式
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")

    # 寫入表頭
    headers = ["文件名", "主題", "主辦單位", "日期", "地點", "積分類別", "原始積分數", "AI初審積分"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment

    # 寫入每個文件的基本信息
    for row, result in enumerate(all_results, start=2):
        ws.cell(row=row, column=1, value=result['文件名'])
        ws.cell(row=row, column=2, value=result['主題'])
        ws.cell(row=row, column=3, value=result['主辦單位'])
        ws.cell(row=row, column=4, value=result['日期'])
        ws.cell(row=row, column=5, value=result['地點'])
        ws.cell(row=row, column=6, value=result['積分類別'])
        ws.cell(row=row, column=7, value=result['原始積分數'])
        ws.cell(row=row, column=8, value=result['AI初審積分'])

    # 創建詳細分析結果工作表
    ws_detail = wb.create_sheet(title="詳細分析結果")
    
    current_row = 1
    for result in all_results:
        # 寫入文件名作為分隔
        ws_detail.cell(row=current_row, column=1, value=f"文件名: {result['文件名']}")
        ws_detail.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        # 寫入基本信息
        basic_info = [
            ("主題", result['主題']),
            ("主辦單位", result['主辦單位']),
            ("日期", result['日期']),
            ("地點", result['地點']),
            ("積分類別", result['積分類別']),
            ("原始積分數", result['原始積分數']),
            ("AI初審積分", result['AI初審積分']),
            ("AI初審積分說明", result['AI初審積分說明'])
        ]
        for info in basic_info:
            ws_detail.cell(row=current_row, column=1, value=info[0])
            ws_detail.cell(row=current_row, column=2, value=info[1])
            current_row += 1

        # 寫入演講主題表格標題
        headers = ["主題", "講者", "主持人", "時間", "持續時間(分鐘)", "AI初審"]
        for col, header in enumerate(headers, start=1):
            cell = ws_detail.cell(row=current_row, column=col, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
        current_row += 1

        # 寫入演講主題詳情
        for topic in result['演講主題']:
            ws_detail.cell(row=current_row, column=1, value=topic['topic'])
            ws_detail.cell(row=current_row, column=2, value=topic['speaker'])
            ws_detail.cell(row=current_row, column=3, value=topic['moderator'])
            ws_detail.cell(row=current_row, column=4, value=topic['time'])
            ws_detail.cell(row=current_row, column=5, value=topic['duration'])
            ws_detail.cell(row=current_row, column=6, value=topic['ai_review'])
            current_row += 1

        # 添加空行作為分隔
        current_row += 2

    # 調整列寬
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)

def process_single_file(client, file_path):
    if not os.path.isfile(file_path):
        print(f"文件不存在：{file_path}")
        return None

    file_extension = os.path.splitext(file_path)[1].lower()
    if file_extension not in ['.pdf', '.doc', '.docx']:
        print(f"不支持的文件類型：{file_extension}")
        return None

    text = analyze_file(file_path)
    if not text:
        return None

    prompt = f"""
    分析以下內容並提取關鍵資訊，以JSON格式返回結果：
    - 主題
    - 主辦單位
    - 日期
    - 地點
    - 演講主題（多筆，每筆包括：主題（包含時間）、講者、主持人、時間（time）、持續時間（分鐘）、AI初審）
    
    JSON格式如下：
    {{
        "主題": "string",
        "主辦單位": "string",
        "日期": "string",
        "地點": "string",
        "積分類別": "string",
        "演講主題": [
            {{
                "topic": "string",
                "speaker": "string",
                "moderator": "string",
                "time": "string",
                "duration": int,
                "ai_review": "string"
            }},
            // ... 可能有多筆
        ]
    }}

    注意：
    1. 主題應包含時間資訊。
    2. QA 或問答時間必須合併到其上方的主題中，不要單獨列出 QA 或問答環節。
    3. 每個主題的時間（time）字段應包括主題開始時間和結束時間，如果有 QA，則結束時間為 QA 結束時間。
       例如："09:00-10:30"或"14:00-15:30（包含QA）"。
    4. 每個主題的持續時間（duration）必須包括該主題的講座時間加上其後的 QA 或問答時間（如果有的話）。
       例如，如果一個主題講座時間為 30 分鐘，之後有 10 分鐘 QA，則該主題的總 duration 應為 40 分鐘。
    5. 如果一個主題包含了 QA 時間，請在該主題的 topic 字段末尾添加 "（包含 QA）"。
    6. Registration, Opening Remarks, Closing Remarks 等項目應包含在演講主題列表中，並視為與會議主題直接相關。
    7. 積分類別計分原則："中華民國糖尿病學會"主辦，或"糖尿病學會"主辦，為甲類，其餘為乙類。
    8. AI初審：針對topic內容進行審查。和糖尿病、高血壓、高血脂或相關併發症有關的，註明"相關"，沒有關係的，註明"不相關"；不確定者，註明"？"。"不相關"者請註明原因。
       Registration, Opening Remarks, Closing Remarks 等項目不需要進行 AI 初審，其時間納入學分總時間。

    以下是分析的內容：

    {text}
    """

    parsed_result = get_gpt4_json_response(client, prompt)

    # 處理主題，重新計算包含 QA 的 duration
    parsed_result['演講主題'] = process_topics(parsed_result['演講主題'])

    # 算總時間和有效時間
    total_duration = sum(topic['duration'] for topic in parsed_result['演講主題'])
    valid_duration = sum(topic['duration'] for topic in parsed_result['演講主題'] 
                         if topic['ai_review'] != "不相關" or 
                         topic['topic'].lower() in ['registration', 'opening remarks', 'closing remarks'])
    
    # 計算積分數
    credit_type = parsed_result['積分類別']
    original_credits = calculate_credits(total_duration, credit_type)
    adjusted_credits = calculate_credits(valid_duration, credit_type)
    
    # 生成 AI 初審積分說明
    ai_review_explanation = f"""
    原始總時數：{total_duration} 分鐘
    原始積分數：{original_credits} 學分（{credit_type}）
    AI 初審後有效時數：{valid_duration} 分鐘
    AI 初審積分：{adjusted_credits} 學分（{credit_type}）
    """
    parsed_result['AI初審積分說明'] = ai_review_explanation.strip()
    
    # 添加積分數和AI初審積分到結果中
    parsed_result['原始積分數'] = original_credits
    parsed_result['AI初審積分'] = adjusted_credits
    
    # 添加文件名到結果中
    parsed_result['文件名'] = os.path.basename(file_path)
    
    return parsed_result

def main():
    client = openai.OpenAI(api_key=OPENAI_API_KEY)

    # 指定包含文件的目錄
    directory = '/home/runner/creditanalyze/'

    # 獲取目錄中所有的 PDF、DOC 和 DOCX 文件
    supported_files = [f for f in os.listdir(directory) if f.lower().endswith(('.pdf', '.doc', '.docx'))]

    all_results = []

    for file in supported_files:
        file_path = os.path.join(directory, file)
        print(f"正在處理文件：{file_path}")

        parsed_result = process_single_file(client, file_path)
        if parsed_result:
            all_results.append(parsed_result)
        else:
            print(f"處理文件 {file_path} 時出錯")

    if all_results:
        # 生成輸出文件名
        output_file = os.path.join(directory, 'combined_analysis_results.xlsx')
        write_to_excel(all_results, output_file)
        print(f"所有結果已輸出到 {output_file}")
    else:
        print("沒有成功處理任何文件")

if __name__ == "__main__":
    main()