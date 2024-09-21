import streamlit as st
import subprocess
import sys

# Add this at the beginning of your script
def install(package):
    subprocess.check_call([sys.executable, "-m", "pip", "install", package])

# Check if cv2 is installed, if not, install it
try:
    import cv2
except ImportError:
    st.info("Installing required packages...")
    install("opencv-python-headless")
    import cv2

# Rest of your imports
import openai
from openai import OpenAI
import os
import json
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from datetime import datetime, timedelta
import textract
import base64
import tempfile
from dotenv import load_dotenv
from PIL import Image
import pytesseract
import PyPDF2
import numpy as np
from pdf2image import convert_from_path
import re
import logging

load_dotenv()  # 載入 .env 檔案中的環境變數

# 設置日誌
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

def preprocess_image(image):
    # 轉換為灰度圖
    gray = cv2.cvtColor(np.array(image), cv2.COLOR_RGB2GRAY)
    
    # 應用自適應閾值
    thresh = cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C, cv2.THRESH_BINARY, 11, 2)
    
    # 去噪
    denoised = cv2.fastNlMeansDenoising(thresh, None, 10, 7, 21)
    
    return Image.fromarray(denoised)

def ocr_image(image):
    preprocessed_image = preprocess_image(image)
    text = pytesseract.image_to_string(preprocessed_image, lang='chi_tra+eng')
    return text

def extract_text_from_pdf_pypdf2(file_path):
    try:
        with open(file_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            text = ""
            for page in reader.pages:
                text += page.extract_text()
        return text
    except Exception as e:
        logger.error(f"PyPDF2 從 PDF 提取文本時出錯: {str(e)}")
        return ""

def extract_text_from_pdf_ocr(file_path):
    try:
        images = convert_from_path(file_path)
        text = ""
        for image in images:
            text += ocr_image(image) + "\n\n"
        return text
    except Exception as e:
        logger.error(f"OCR 從 PDF 提取文本時出錯: {str(e)}")
        return ""

def extract_text_from_pdf_gpt4_vision(client, file_path):
    try:
        images = convert_from_path(file_path)
        all_text = ""
        for i, image in enumerate(images):
            # 保存圖片到臨時文件
            with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as temp_file:
                image.save(temp_file, format="PNG")
                temp_file_path = temp_file.name

            # 讀取圖片文件
            with open(temp_file_path, "rb") as image_file:
                base64_image = base64.b64encode(image_file.read()).decode('utf-8')

            response = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": "請閱讀這張圖片，並提取所有可見的文字內容。只需返回提取的文字，不需要任何解釋或格式化。"},
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}"
                                }
                            },
                        ],
                    }
                ],
                max_tokens=1000,
            )
            
            all_text += response.choices[0].message.content + "\n\n"
            
            # 刪除臨時文件
            os.unlink(temp_file_path)

        return all_text.strip()
    except Exception as e:
        logger.error(f"GPT-4 Vision 從 PDF 提取文本時出錯: {str(e)}")
        return ""

def extract_text_from_pdf(client, file_path):
    # 首先嘗試使用 PyPDF2
    text = extract_text_from_pdf_pypdf2(file_path)  # 注意這裡需要傳入 client
    
    # 如果 PyPDF2 提取的文本為空或太短，則使用 GPT-4 Vision
    if len(text.strip()) < 100:
        logger.info("PyPDF2 提取的文本不足，切換到 GPT-4 Vision")
        text = extract_text_from_pdf_gpt4_vision(client, file_path)
    
    return text

def clean_text(text):
    # 移除多餘的空白字符
    text = re.sub(r'\s+', ' ', text)
    # 移除非打印字符
    text = ''.join(char for char in text if char.isprintable() or char.isspace())
    return text.strip()

def analyze_file(client, file_path):
    file_extension = os.path.splitext(file_path)[1].lower()
    try:
        if file_extension == '.pdf':
            text = extract_text_from_pdf(client, file_path)  # 注意這裡需要傳入 client
        elif file_extension in ['.doc', '.docx']:
            text = textract.process(file_path).decode('utf-8', errors='ignore')
        elif file_extension in ['.jpg', '.jpeg', '.png']:
            image = Image.open(file_path)
            text = ocr_image(image)
        else:
            logger.error(f"不支援的檔案類型：{file_extension}")
            return None

        text = clean_text(text)
        
        if not text.strip():
            logger.warning(f"無法從檔案 {file_path} 提取文本")
            return None

        return text
    except Exception as e:
        logger.error(f"處理檔案 {file_path} 時出錯: {str(e)}")
        return None

def calculate_credits(duration_minutes, credit_type):
    if credit_type == "甲類":
        return math.ceil(duration_minutes / 50)
    else:  # 乙類
        return round(duration_minutes / 50 * 0.5, 1)

def get_gpt4_json_response(client, prompt):
    try:
        response = client.chat.completions.create(
            model="gpt-4o",
            messages=[
                {"role": "system", "content": "你是一個專業的文件分析助手。請以JSON格式返回分析結果。"},
                {"role": "user", "content": prompt}
            ],
            response_format={"type": "json_object"},
            max_tokens=1500  # 增加 token 限制
        )
        return json.loads(response.choices[0].message.content)
    except Exception as e:
        logger.error(f"GPT-4 API 調用失敗: {str(e)}")
        return None

def calculate_duration(start_time, end_time):
    start = datetime.strptime(start_time, "%H:%M")
    end = datetime.strptime(end_time, "%H:%M")
    if end < start:
        end += timedelta(days=1)  # 處理跨午夜的情況
    duration = end - start
    return int(duration.total_seconds() / 60)  # 返回分鐘數

def process_topics(topics):
    processed_topics = []
    for topic in topics:
        processed_topic = {
            "topic": topic.get("topic", ""),
            "speaker": topic.get("speaker", ""),  # 保留為空字串如果沒有講者
            "moderator": topic.get("moderator", ""),
            "time": topic.get("time", ""),
            "duration": int(topic.get("duration", 0)),
            "ai_review": topic.get("ai_review", "")
        }
        processed_topics.append(processed_topic)
    return processed_topics

def write_to_excel(all_results, output_file):
    wb = Workbook()
    ws = wb.active
    ws.title = "課程分析結果"

    # 設置標題樣式
    title_font = Font(bold=True)
    title_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    title_alignment = Alignment(horizontal="center", vertical="center")

    # 寫入表頭
    headers = ["文件名", "主", "單位", "日期", "點", "積分類別", "原積分數", "AI初審積分"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = title_font
        cell.fill = title_fill
        cell.alignment = title_alignment

    # 寫入每個文件的基本信息
    for row, result in enumerate(all_results, start=2):
        ws.cell(row=row, column=1, value=result.get('文件名', ''))
        ws.cell(row=row, column=2, value=result.get('主題', ''))
        ws.cell(row=row, column=3, value=result.get('主辦單位', ''))
        ws.cell(row=row, column=4, value=result.get('日期', ''))
        ws.cell(row=row, column=5, value=result.get('地點', ''))
        ws.cell(row=row, column=6, value=result.get('積分類別', ''))
        ws.cell(row=row, column=7, value=result.get('原始積分數', 0))
        ws.cell(row=row, column=8, value=result.get('AI初審積分', 0))

    # 創建詳細分析結果工作表
    ws_detail = wb.create_sheet(title="詳細分析結果")
    
    current_row = 1
    for result in all_results:
        # 寫入文件名作為分隔
        ws_detail.cell(row=current_row, column=1, value=f"文件名: {result.get('文件名', '')}")
        ws_detail.cell(row=current_row, column=1).font = Font(bold=True)
        current_row += 1

        # 寫入基本信息
        basic_info = [
            ("主題", result.get('主題', '')),
            ("主辦單位", result.get('主辦單位', '')),
            ("日期", result.get('日期', '')),
            ("地點", result.get('地點', '')),
            ("積分類別", result.get('積分類別', '')),
            ("原始積分數", result.get('原始積分數', 0)),
            ("AI初審積分", result.get('AI初審積分', 0)),
            ("AI初審積分說明", result.get('AI初審積分說明', ''))
        ]
        for info in basic_info:
            ws_detail.cell(row=current_row, column=1, value=info[0])
            ws_detail.cell(row=current_row, column=2, value=info[1])
            current_row += 1

        # 寫入演講主題表格題
        headers = ["主題", "講者", "主持人", "時間", "持續時間(分鐘)", "AI初審"]
        for col, header in enumerate(headers, start=1):
            cell = ws_detail.cell(row=current_row, column=col, value=header)
            cell.font = title_font
            cell.fill = title_fill
            cell.alignment = title_alignment
        current_row += 1

        # 寫入演講主題詳情
        for topic in result.get('演講主題', []):
            ws_detail.cell(row=current_row, column=1, value=topic.get('topic', ''))
            ws_detail.cell(row=current_row, column=2, value=topic.get('speaker', ''))
            ws_detail.cell(row=current_row, column=3, value=topic.get('moderator', ''))
            ws_detail.cell(row=current_row, column=4, value=topic.get('time', ''))
            ws_detail.cell(row=current_row, column=5, value=topic.get('duration', 0))
            ws_detail.cell(row=current_row, column=6, value=topic.get('ai_review', ''))
            current_row += 1

        # 添加空行作為分隔
        current_row += 2

    # 調整列寬
    for ws in wb.worksheets:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_file)

def is_special_item(topic):
    special_keywords = [
        'registration', '報到', '簽到',
        'opening', '開幕', '開場',
        'closing', '閉幕', '結束',
        'panel discussion', '座談會', '討論會',
        'break', '休息',
        'lunch', '午餐',
        'dinner', '晚餐',
        'welcome', '歡迎'
    ]
    topic_lower = topic.lower()
    return any(keyword in topic_lower for keyword in special_keywords)

def process_single_file(client, file_path):
    try:
        analyzed_content = analyze_file(client, file_path)  # 注意這裡需要傳入 client
        if not analyzed_content:
            return None

        prompt = f"""
        分析以下內容並提取關鍵資訊，以JSON格式返回結果：
        - 主題
        - 主辦單位
        - 日期
        - 地點
        - 演講主題（多筆，每筆包括：主題（包含時間）、講者、主持人、時間（time）、持續時間（分鐘）、AI初審）
        
        JSON格式如下：
        {{
            "主題": "string",
            "主辦單位": "string",
            "日期": "string",
            "地點": "string",
            "積分類別": "string",
            "演講主題": [
                {{
                    "topic": "string",
                    "speaker": "string",
                    "moderator": "string",
                    "time": "string",
                    "duration": int,
                    "ai_review": "string"
                }},
                // ... 可能有多筆
            ]
        }}

        注意：
        1. 主題應包含時間資訊。
        2. QA 或問答時間必須合併到上方的主題中，不要單獨列出 QA 或問答環節。
        3. 每個主題的時間（time）字段應包括主題開始時間和結束時間，如果有 QA，則結束時間為 QA 結束時間。
           例如："09:00-10:30"或"14:00-15:30（包含QA）"。
        4. 每個主題的持續時間（duration）必須包括該主題的講座時間加上其後的 QA 或問答時間（如果有的話）。
           例如，如果一個主題講座時間為 30 分鐘，之後有 10 分鐘 QA，則該主題的總 duration 應為 40 分鐘。
        5. 如果一個主題包含了 QA 時間，請在該主題的 topic 字段末尾添加 "（包含 QA）"。
        6. Registration, Opening Remarks, Closing Remarks 等項目應包含在演講主題列表中，並視為與會議主題直接相關。
        7. 積分類別計分原則："中華民國糖尿病學會"主辦，或"糖尿病學會"主辦，為甲類，其餘為乙類。
        8. AI初審：針對topic內容進行審查。和糖尿病、高血壓、高血脂或相關併發症有關的，註明"相關"，沒有關係的，註明"不相關"；不確定者，註明"？"。"不相關"者請註明原因。
           
        9. 有以下相關字眼，Registration, Opening Remarks, Closing Remarks, Pannel Discussion, 等項目不需要進行 AI 初審，不需列是否相關，但其時間應入學分總時間。
        10. 如果無法辨識講者名字，請將講者欄位留空。

        以下是需要分析的內容：
        {analyzed_content}
        """

        parsed_result = get_gpt4_json_response(client, prompt)
        if not parsed_result:
            logger.error("GPT-4 分析失敗")
            return None

        # 確保 '演講主題' 字段存在且為列表
        if '演講主題' not in parsed_result or not isinstance(parsed_result['演講主題'], list):
            logger.warning(f"警告：分析結果中沒有有效的 '演講主題' 字段")
            parsed_result['演講主題'] = []

        # 處理主題，重新計算包含 QA 的 duration
        parsed_result['演講主題'] = process_topics(parsed_result['演講主題'])

        # 算總時間和有效時間
        total_duration = 0
        valid_duration = 0
        for topic in parsed_result['演講主題']:
            duration = topic.get('duration', 0)
            total_duration += duration
            
            # 檢查是否為特殊項目
            if is_special_item(topic.get('topic', '')):
                valid_duration += duration
                # 確保這些項目不進行 AI 初審
                topic['ai_review'] = ''
            else:
                # 檢查 AI 初審結果
                ai_review = topic.get('ai_review', '').lower()
                if ai_review == '相關' or ai_review == '':
                    valid_duration += duration
                elif '?' in ai_review:
                    # 對於不確定的情況，計入一半的時間
                    valid_duration += duration / 2

        # 計算積分
        credit_type = parsed_result.get('積分類別', '')
        credits = calculate_credits(valid_duration, credit_type)

        # 添加計算結果到 parsed_result
        parsed_result['原始積分數'] = calculate_credits(total_duration, credit_type)
        parsed_result['AI初審積分'] = credits
        parsed_result['AI初審積分說明'] = f"有效時間：{valid_duration} 分鐘，總時間：{total_duration} 分鐘"

        return parsed_result
    except Exception as e:
        logger.error(f"處理檔案時發生異常: {str(e)}")
        return None

def main():
    st.title("糖尿病學會 學分分析助手")

    # 添加文件格式說明
    st.markdown("""
    **注意：** 
    - 支援的文件格式：PDF, DOCX, JPG, JPEG, PNG
    - 不支援 .doc 格式，請先將 .doc 文件轉換為 .docx 格式後再上傳
    - 轉換方法：使用 Microsoft Word 打開 .doc 文件，然後「另存新檔」為 .docx 格式
    - 尚不支援複雜PDF文件的處理
    """)

    # 在側邊欄中設置 API 金鑰輸入
    st.sidebar.title("設定")
    openai_api_key = st.sidebar.text_input(
        label="請輸入您的 OpenAI API 金鑰：",
        type='password',
        placeholder="例如：sk-2twmA88un4...",
        help="您可以從 https://platform.openai.com/account/api-keys/ 獲取您的 API 金鑰"
    )

    # 添加空白空間，將聯絡信息推到底部
    st.sidebar.empty()
    st.sidebar.empty()
    st.sidebar.empty()
    
    # 在側邊欄底部添加聯絡信息
    st.sidebar.markdown("---")  # 添加分隔線
    st.sidebar.markdown("**程式設計:** Tseng Yao Hsien \n Tung's Taichung Metroharbor Hospital")
    st.sidebar.markdown("**聯絡:** LINE: zinojeng")

    # 檢查是否輸入了 API 金鑰
    if not openai_api_key:
        st.warning("請在側邊欄輸入有效的 OpenAI API 金鑰以繼續")
        return

    # 使用 API 金鑰，但不存儲它
    client = openai.OpenAI(api_key=openai_api_key)

    # 檔案上傳
    uploaded_files = st.file_uploader("上傳檔案（支援 PDF, DOCX, JPG, JPEG, PNG）", accept_multiple_files=True, type=['pdf', 'docx', 'jpg', 'jpeg', 'png'])

    if uploaded_files:
        all_results = []
        progress_bar = st.progress(0)

        for index, uploaded_file in enumerate(uploaded_files):
            original_filename = uploaded_file.name
            st.write(f"正在處理檔案：{original_filename}")

            with tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(original_filename)[1]) as temp_file:
                temp_file.write(uploaded_file.getvalue())
                temp_file_path = temp_file.name

            try:
                parsed_result = process_single_file(client, temp_file_path)
                if parsed_result:
                    parsed_result['文件名'] = original_filename  # 使用原始檔案名
                    all_results.append(parsed_result)
                    st.success(f"成功處理檔案：{original_filename}")
                    
                    # 顯示辨識結
                    #st.subheader(f"文件：{original_filename} 的辨識結果")
                    #st.json(parsed_result)

                    # 檢查需要人工審核的項目
                    need_review = [topic for topic in parsed_result['演講主題'] if not topic['speaker'] or topic['ai_review'] == '？']
                    if need_review:
                        st.warning("以下項目可能需要人工審核：")
                        for topic in need_review:
                            st.write(f"- 演講主題: {topic['topic']}, 講者: {'未識別' if not topic['speaker'] else topic['speaker']}")
                else:
                    st.error(f"處理檔案 {original_filename} 時出錯")
            except Exception as e:
                logger.error(f"理檔案 {original_filename} 時發生異常: {str(e)}")
                st.error(f"處理檔案 {original_filename} 時發生異常: {str(e)}")
            finally:
                os.unlink(temp_file_path)
            
            # 更新進度條
            progress_bar.progress((index + 1) / len(uploaded_files))

        if all_results:
            output_file = '分析結果總表.xlsx'
            write_to_excel(all_results, output_file)
            st.success(f"所有結果已輸出到 {output_file}")
            
            with open(output_file, "rb") as file:
                st.download_button(
                    label="下載分析結果",
                    data=file,
                    file_name=output_file,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
        else:
            st.warning("沒有成功處理任何檔案")

def determine_credit_type(organizer):
    # 將輸入轉換為小寫並去除首尾空格，以便進行不區分大小寫的比較
    organizer = organizer.lower().strip()
    
    # 定義甲類積分的主辦單位列表
    class_a_organizers = [
        "中華民國糖尿病學會",
        "糖尿病學會",
        "中華民國內分泌學會",
        "內分泌學會"
    ]
    
    # 檢查是否為甲類積分主辦單位
    for org in class_a_organizers:
        if org.lower() == organizer or organizer.endswith(org.lower()):
            return "甲類"
    
    # 如果不是甲類，則返回乙類
    return "乙類"

# 在解析結果時使用
parsed_result['積分類別'] = determine_credit_type(parsed_result.get('主辦單位', ''))

# 添加日誌記錄
import logging
logging.info(f"主辦單位: {parsed_result.get('主辦單位', '')}, 判斷積分類別: {parsed_result['積分類別']}")

if __name__ == "__main__":
    main()  # 這會在文件末尾添加一個空行
